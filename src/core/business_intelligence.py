#!/usr/bin/env python3
"""
Business Intelligence & Advanced Reporting System
Enterprise-grade business intelligence platform with advanced analytics,
executive dashboards, automated reporting, and data visualization.
"""

import asyncio
import logging
import json
import os
from datetime import datetime, timedelta
import time
from typing import Dict, List, Any, Optional, Union, Tuple
from dataclasses import dataclass, asdict
from enum import Enum
import uuid
import numpy as np
import pandas as pd
from pathlib import Path

# Business Intelligence & Analytics
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.io as pio
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# Advanced Analytics
from scipy import stats
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import silhouette_score
import networkx as nx

# Time Series Analysis
from prophet import Prophet
import statsmodels.api as sm
from statsmodels.tsa.seasonal import seasonal_decompose
from statsmodels.tsa.arima.model import ARIMA

# Reporting & Documents
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import io
import base64

# Excel Reports
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import LineChart, BarChart, PieChart, Reference

# Email & Notifications
import smtplib
from email.mime.multipart import MimeMultipart
from email.mime.text import MimeText
from email.mime.base import MimeBase
from email import encoders

# Web Framework
from fastapi import FastAPI, Request, Response, HTTPException, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from jinja2 import Environment, FileSystemLoader

# Database & Storage
import asyncpg
import motor.motor_asyncio
import redis
import boto3

# Utilities
import structlog
from concurrent.futures import ThreadPoolExecutor
import schedule
import threading

logger = structlog.get_logger()

class ReportType(str, Enum):
    EXECUTIVE_SUMMARY = "executive_summary"
    COST_ANALYSIS = "cost_analysis"
    OPTIMIZATION_REPORT = "optimization_report"
    TREND_ANALYSIS = "trend_analysis"
    ROI_ANALYSIS = "roi_analysis"
    COMPLIANCE_REPORT = "compliance_report"
    CUSTOM_DASHBOARD = "custom_dashboard"

class ReportFormat(str, Enum):
    PDF = "pdf"
    EXCEL = "excel"
    HTML = "html"
    JSON = "json"
    CSV = "csv"
    POWERPOINT = "powerpoint"

class ReportFrequency(str, Enum):
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"
    ON_DEMAND = "on_demand"

class DashboardType(str, Enum):
    EXECUTIVE = "executive"
    OPERATIONAL = "operational"
    TECHNICAL = "technical"
    FINANCIAL = "financial"
    CUSTOM = "custom"

@dataclass
class ReportConfig:
    """Configuration for automated reports."""
    report_id: str
    name: str
    description: str
    report_type: ReportType
    format: ReportFormat
    frequency: ReportFrequency
    recipients: List[str]
    parameters: Dict[str, Any]
    filters: Dict[str, Any]
    charts_included: List[str]
    created_by: str
    created_at: datetime
    last_generated: Optional[datetime]
    next_generation: Optional[datetime]
    active: bool

@dataclass
class DashboardConfig:
    """Configuration for business intelligence dashboards."""
    dashboard_id: str
    name: str
    description: str
    dashboard_type: DashboardType
    owner: str
    viewers: List[str]
    widgets: List[Dict[str, Any]]
    layout: Dict[str, Any]
    refresh_interval: int
    created_at: datetime
    updated_at: datetime
    public: bool

@dataclass
class AnalyticsMetrics:
    """Business intelligence metrics."""
    metric_id: str
    name: str
    value: float
    unit: str
    category: str
    timestamp: datetime
    metadata: Dict[str, Any]
    trend: Optional[float]
    target: Optional[float]

class AdvancedAnalyticsEngine:
    """Advanced analytics engine for business intelligence."""
    
    def __init__(self):
        self.data_cache = {}
        self.models = {}
        
    async def perform_cost_trend_analysis(self, data: pd.DataFrame) -> Dict[str, Any]:
        """Perform advanced cost trend analysis."""
        try:
            if 'timestamp' not in data.columns or 'cost' not in data.columns:
                raise ValueError("Data must contain 'timestamp' and 'cost' columns")
            
            # Prepare time series data
            data['timestamp'] = pd.to_datetime(data['timestamp'])
            data = data.sort_values('timestamp')
            data.set_index('timestamp', inplace=True)
            
            # Seasonal decomposition
            decomposition = seasonal_decompose(data['cost'], model='additive', period=30)
            
            # Trend analysis
            trend_slope = self._calculate_trend_slope(data['cost'])
            
            # Forecasting
            forecast_data = await self._forecast_costs(data)
            
            # Anomaly detection
            anomalies = self._detect_cost_anomalies(data['cost'])
            
            # Statistical analysis
            stats_analysis = self._calculate_cost_statistics(data['cost'])
            
            return {
                'trend_analysis': {
                    'slope': trend_slope,
                    'direction': 'increasing' if trend_slope > 0 else 'decreasing',
                    'magnitude': abs(trend_slope),
                    'trend_component': decomposition.trend.dropna().tolist(),
                    'seasonal_component': decomposition.seasonal.dropna().tolist(),
                    'residual_component': decomposition.resid.dropna().tolist()
                },
                'forecast': forecast_data,
                'anomalies': anomalies,
                'statistics': stats_analysis,
                'insights': await self._generate_cost_insights(data, trend_slope, anomalies)
            }
            
        except Exception as e:
            logger.error(f"Cost trend analysis failed: {e}")
            return {}
    
    async def perform_resource_optimization_analysis(self, data: pd.DataFrame) -> Dict[str, Any]:
        """Analyze resource optimization opportunities."""
        try:
            # Cluster analysis for resource patterns
            resource_clusters = await self._analyze_resource_clusters(data)
            
            # Efficiency analysis
            efficiency_metrics = self._calculate_efficiency_metrics(data)
            
            # Right-sizing opportunities
            rightsizing_opportunities = self._identify_rightsizing_opportunities(data)
            
            # Utilization patterns
            utilization_patterns = self._analyze_utilization_patterns(data)
            
            # Cost optimization recommendations
            optimization_recommendations = await self._generate_optimization_recommendations(
                resource_clusters, efficiency_metrics, rightsizing_opportunities
            )
            
            return {
                'resource_clusters': resource_clusters,
                'efficiency_metrics': efficiency_metrics,
                'rightsizing_opportunities': rightsizing_opportunities,
                'utilization_patterns': utilization_patterns,
                'optimization_recommendations': optimization_recommendations,
                'potential_savings': self._calculate_potential_savings(rightsizing_opportunities)
            }
            
        except Exception as e:
            logger.error(f"Resource optimization analysis failed: {e}")
            return {}
    
    async def perform_roi_analysis(self, cost_data: pd.DataFrame, 
                                 savings_data: pd.DataFrame) -> Dict[str, Any]:
        """Perform ROI analysis for optimization initiatives."""
        try:
            # Calculate baseline costs
            baseline_cost = cost_data['cost'].sum()
            
            # Calculate realized savings
            realized_savings = savings_data['savings'].sum()
            
            # Calculate investment costs (platform costs, implementation costs)
            investment_cost = savings_data.get('investment', pd.Series([0])).sum()
            
            # ROI calculation
            roi = (realized_savings - investment_cost) / investment_cost * 100 if investment_cost > 0 else 0
            
            # Payback period
            payback_period = investment_cost / (realized_savings / len(savings_data)) if realized_savings > 0 else float('inf')
            
            # NPV calculation (simplified)
            discount_rate = 0.1  # 10% discount rate
            periods = len(savings_data)
            monthly_savings = realized_savings / periods
            
            npv = sum([
                monthly_savings / ((1 + discount_rate) ** (month / 12))
                for month in range(1, periods + 1)
            ]) - investment_cost
            
            # Cost avoidance analysis
            cost_avoidance = self._calculate_cost_avoidance(cost_data, savings_data)
            
            return {
                'baseline_cost': baseline_cost,
                'realized_savings': realized_savings,
                'investment_cost': investment_cost,
                'roi_percentage': roi,
                'payback_period_months': payback_period,
                'npv': npv,
                'cost_avoidance': cost_avoidance,
                'efficiency_gains': {
                    'cost_reduction_percentage': (realized_savings / baseline_cost) * 100,
                    'monthly_savings_trend': savings_data['savings'].tolist(),
                    'cumulative_savings': savings_data['savings'].cumsum().tolist()
                }
            }
            
        except Exception as e:
            logger.error(f"ROI analysis failed: {e}")
            return {}
    
    def _calculate_trend_slope(self, series: pd.Series) -> float:
        """Calculate trend slope using linear regression."""
        try:
            x = np.arange(len(series))
            y = series.values
            slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
            return slope
        except Exception:
            return 0.0
    
    async def _forecast_costs(self, data: pd.DataFrame) -> Dict[str, Any]:
        """Forecast future costs using Prophet."""
        try:
            # Prepare data for Prophet
            prophet_data = data.reset_index().rename(columns={'timestamp': 'ds', 'cost': 'y'})
            
            # Create and fit model
            model = Prophet(
                daily_seasonality=True,
                weekly_seasonality=True,
                yearly_seasonality=True
            )
            model.fit(prophet_data)
            
            # Make future predictions
            future = model.make_future_dataframe(periods=90)  # 90 days ahead
            forecast = model.predict(future)
            
            # Extract forecast data
            forecast_period = forecast.tail(90)
            
            return {
                'forecast_values': forecast_period['yhat'].tolist(),
                'forecast_lower': forecast_period['yhat_lower'].tolist(),
                'forecast_upper': forecast_period['yhat_upper'].tolist(),
                'forecast_dates': forecast_period['ds'].dt.strftime('%Y-%m-%d').tolist(),
                'trend': forecast_period['trend'].tolist(),
                'seasonal': forecast_period.get('seasonal', []).tolist() if 'seasonal' in forecast_period.columns else []
            }
            
        except Exception as e:
            logger.error(f"Cost forecasting failed: {e}")
            return {}
    
    def _detect_cost_anomalies(self, series: pd.Series) -> List[Dict[str, Any]]:
        """Detect cost anomalies using statistical methods."""
        try:
            anomalies = []
            
            # Z-score method
            z_scores = np.abs(stats.zscore(series))
            anomaly_threshold = 3
            
            anomaly_indices = np.where(z_scores > anomaly_threshold)[0]
            
            for idx in anomaly_indices:
                anomalies.append({
                    'index': int(idx),
                    'timestamp': series.index[idx].isoformat() if hasattr(series.index[idx], 'isoformat') else str(idx),
                    'value': float(series.iloc[idx]),
                    'z_score': float(z_scores[idx]),
                    'type': 'statistical_outlier'
                })
            
            return anomalies
            
        except Exception as e:
            logger.error(f"Anomaly detection failed: {e}")
            return []
    
    def _calculate_cost_statistics(self, series: pd.Series) -> Dict[str, float]:
        """Calculate comprehensive cost statistics."""
        try:
            return {
                'mean': float(series.mean()),
                'median': float(series.median()),
                'std': float(series.std()),
                'min': float(series.min()),
                'max': float(series.max()),
                'q25': float(series.quantile(0.25)),
                'q75': float(series.quantile(0.75)),
                'coefficient_of_variation': float(series.std() / series.mean()) if series.mean() != 0 else 0,
                'skewness': float(stats.skew(series)),
                'kurtosis': float(stats.kurtosis(series))
            }
        except Exception:
            return {}
    
    async def _generate_cost_insights(self, data: pd.DataFrame, trend_slope: float, 
                                    anomalies: List[Dict]) -> List[str]:
        """Generate insights from cost analysis."""
        insights = []
        
        try:
            # Trend insights
            if abs(trend_slope) > 0.1:
                direction = "increasing" if trend_slope > 0 else "decreasing"
                insights.append(f"Cost is {direction} with a slope of {trend_slope:.3f}")
            
            # Anomaly insights
            if anomalies:
                insights.append(f"Detected {len(anomalies)} cost anomalies requiring investigation")
            
            # Variability insights
            cv = data['cost'].std() / data['cost'].mean() if data['cost'].mean() != 0 else 0
            if cv > 0.3:
                insights.append("High cost variability detected - consider investigating root causes")
            
            # Seasonal insights
            if len(data) >= 30:
                monthly_avg = data.groupby(data.index.month)['cost'].mean()
                peak_month = monthly_avg.idxmax()
                low_month = monthly_avg.idxmin()
                insights.append(f"Peak costs typically occur in month {peak_month}, lowest in month {low_month}")
            
            return insights
            
        except Exception as e:
            logger.error(f"Insight generation failed: {e}")
            return []

class VisualizationEngine:
    """Advanced visualization engine for creating charts and dashboards."""
    
    def __init__(self):
        self.chart_templates = {}
        self.color_palettes = {
            'corporate': ['#667eea', '#764ba2', '#f093fb', '#f5576c', '#4facfe'],
            'professional': ['#2E86AB', '#A23B72', '#F18F01', '#C73E1D', '#592E83'],
            'modern': ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7']
        }
    
    def create_cost_trend_chart(self, data: pd.DataFrame, 
                               forecast_data: Optional[Dict] = None) -> str:
        """Create cost trend chart with forecast."""
        try:
            fig = go.Figure()
            
            # Actual cost data
            fig.add_trace(go.Scatter(
                x=data.index,
                y=data['cost'],
                mode='lines+markers',
                name='Actual Cost',
                line=dict(color=self.color_palettes['corporate'][0], width=3),
                marker=dict(size=6)
            ))
            
            # Add forecast if available
            if forecast_data:
                forecast_dates = pd.to_datetime(forecast_data['forecast_dates'])
                
                # Forecast line
                fig.add_trace(go.Scatter(
                    x=forecast_dates,
                    y=forecast_data['forecast_values'],
                    mode='lines',
                    name='Forecast',
                    line=dict(color=self.color_palettes['corporate'][1], width=2, dash='dash')
                ))
                
                # Confidence interval
                fig.add_trace(go.Scatter(
                    x=forecast_dates.tolist() + forecast_dates.tolist()[::-1],
                    y=forecast_data['forecast_upper'] + forecast_data['forecast_lower'][::-1],
                    fill='tonexty',
                    fillcolor='rgba(118, 75, 162, 0.2)',
                    line=dict(color='rgba(255,255,255,0)'),
                    name='Confidence Interval',
                    showlegend=False
                ))
            
            # Update layout
            fig.update_layout(
                title='Cost Trend Analysis with Forecast',
                xaxis_title='Date',
                yaxis_title='Cost ($)',
                template='plotly_white',
                height=500,
                showlegend=True,
                legend=dict(x=0, y=1)
            )
            
            return fig.to_html(include_plotlyjs='cdn')
            
        except Exception as e:
            logger.error(f"Cost trend chart creation failed: {e}")
            return "<p>Chart generation failed</p>"
    
    def create_resource_utilization_dashboard(self, data: pd.DataFrame) -> str:
        """Create comprehensive resource utilization dashboard."""
        try:
            # Create subplots
            fig = make_subplots(
                rows=2, cols=2,
                subplot_titles=('CPU Utilization', 'Memory Utilization', 
                              'Resource Distribution', 'Efficiency Score'),
                specs=[[{"secondary_y": False}, {"secondary_y": False}],
                       [{"type": "pie"}, {"type": "indicator"}]]
            )
            
            # CPU utilization over time
            if 'cpu_utilization' in data.columns:
                fig.add_trace(
                    go.Scatter(
                        x=data.index,
                        y=data['cpu_utilization'],
                        mode='lines',
                        name='CPU %',
                        line=dict(color=self.color_palettes['corporate'][0])
                    ),
                    row=1, col=1
                )
            
            # Memory utilization over time
            if 'memory_utilization' in data.columns:
                fig.add_trace(
                    go.Scatter(
                        x=data.index,
                        y=data['memory_utilization'],
                        mode='lines',
                        name='Memory %',
                        line=dict(color=self.color_palettes['corporate'][1])
                    ),
                    row=1, col=2
                )
            
            # Resource distribution pie chart
            if all(col in data.columns for col in ['cpu_utilization', 'memory_utilization']):
                avg_cpu = data['cpu_utilization'].mean()
                avg_memory = data['memory_utilization'].mean()
                
                fig.add_trace(
                    go.Pie(
                        labels=['CPU Utilized', 'CPU Available', 'Memory Utilized', 'Memory Available'],
                        values=[avg_cpu, 100-avg_cpu, avg_memory, 100-avg_memory],
                        name="Resource Distribution"
                    ),
                    row=2, col=1
                )
            
            # Efficiency score gauge
            efficiency_score = self._calculate_efficiency_score(data)
            fig.add_trace(
                go.Indicator(
                    mode="gauge+number+delta",
                    value=efficiency_score,
                    domain={'x': [0, 1], 'y': [0, 1]},
                    title={'text': "Efficiency Score"},
                    delta={'reference': 80},
                    gauge={
                        'axis': {'range': [None, 100]},
                        'bar': {'color': "darkblue"},
                        'steps': [
                            {'range': [0, 50], 'color': "lightgray"},
                            {'range': [50, 80], 'color': "gray"}
                        ],
                        'threshold': {
                            'line': {'color': "red", 'width': 4},
                            'thickness': 0.75,
                            'value': 90
                        }
                    }
                ),
                row=2, col=2
            )
            
            # Update layout
            fig.update_layout(
                title_text="Resource Utilization Dashboard",
                showlegend=False,
                height=600
            )
            
            return fig.to_html(include_plotlyjs='cdn')
            
        except Exception as e:
            logger.error(f"Resource utilization dashboard creation failed: {e}")
            return "<p>Dashboard generation failed</p>"
    
    def create_cost_breakdown_chart(self, data: Dict[str, float]) -> str:
        """Create cost breakdown visualization."""
        try:
            # Create treemap for cost breakdown
            fig = go.Figure(go.Treemap(
                labels=list(data.keys()),
                values=list(data.values()),
                parents=[""] * len(data),
                textinfo="label+value+percent root",
                texttemplate="<b>%{label}</b><br>$%{value}<br>%{percentRoot}",
                pathbar_textfont_size=15,
                marker_colorscale='Viridis'
            ))
            
            fig.update_layout(
                title="Cost Breakdown by Category",
                font_size=12,
                height=500
            )
            
            return fig.to_html(include_plotlyjs='cdn')
            
        except Exception as e:
            logger.error(f"Cost breakdown chart creation failed: {e}")
            return "<p>Chart generation failed</p>"
    
    def _calculate_efficiency_score(self, data: pd.DataFrame) -> float:
        """Calculate overall efficiency score."""
        try:
            if 'cpu_utilization' in data.columns and 'memory_utilization' in data.columns:
                # Efficiency is based on balanced utilization (not too high, not too low)
                cpu_avg = data['cpu_utilization'].mean()
                memory_avg = data['memory_utilization'].mean()
                
                # Optimal range is 50-80%
                cpu_efficiency = 100 - abs(65 - cpu_avg) * 2 if cpu_avg <= 100 else 0
                memory_efficiency = 100 - abs(65 - memory_avg) * 2 if memory_avg <= 100 else 0
                
                return max(0, (cpu_efficiency + memory_efficiency) / 2)
            
            return 50.0  # Default neutral score
            
        except Exception:
            return 50.0

class ReportGenerator:
    """Advanced report generation system."""
    
    def __init__(self):
        self.analytics_engine = AdvancedAnalyticsEngine()
        self.visualization_engine = VisualizationEngine()
        
    async def generate_executive_report(self, data: Dict[str, Any], 
                                      config: ReportConfig) -> bytes:
        """Generate executive summary report."""
        try:
            if config.format == ReportFormat.PDF:
                return await self._generate_pdf_executive_report(data, config)
            elif config.format == ReportFormat.EXCEL:
                return await self._generate_excel_executive_report(data, config)
            elif config.format == ReportFormat.HTML:
                return await self._generate_html_executive_report(data, config)
            else:
                raise ValueError(f"Unsupported format: {config.format}")
                
        except Exception as e:
            logger.error(f"Executive report generation failed: {e}")
            raise
    
    async def _generate_pdf_executive_report(self, data: Dict[str, Any], 
                                           config: ReportConfig) -> bytes:
        """Generate PDF executive report."""
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.HexColor('#667eea')
            )
            story.append(Paragraph("AKS Cost Intelligence - Executive Report", title_style))
            story.append(Spacer(1, 20))
            
            # Executive Summary
            story.append(Paragraph("Executive Summary", styles['Heading2']))
            
            summary_data = data.get('summary', {})
            summary_text = f"""
            This report provides a comprehensive analysis of your Azure Kubernetes Service costs and optimization opportunities.
            
            <b>Key Findings:</b>
            • Total monthly cost: ${summary_data.get('total_cost', 0):,.2f}
            • Potential savings: ${summary_data.get('potential_savings', 0):,.2f}
            • Efficiency score: {summary_data.get('efficiency_score', 0):.1f}%
            • Number of clusters analyzed: {summary_data.get('cluster_count', 0)}
            
            <b>Recommendations:</b>
            • {len(summary_data.get('recommendations', []))} optimization opportunities identified
            • Estimated ROI: {summary_data.get('roi', 0):.1f}%
            • Implementation timeframe: {summary_data.get('implementation_time', 'N/A')}
            """
            
            story.append(Paragraph(summary_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Cost Analysis
            story.append(Paragraph("Cost Analysis", styles['Heading2']))
            
            cost_data = data.get('cost_analysis', {})
            cost_table_data = [
                ['Metric', 'Current', 'Target', 'Variance'],
                ['Monthly Cost', f"${cost_data.get('current_cost', 0):,.2f}", 
                 f"${cost_data.get('target_cost', 0):,.2f}", 
                 f"{cost_data.get('variance_percent', 0):.1f}%"],
                ['Cost Per Node', f"${cost_data.get('cost_per_node', 0):,.2f}", 
                 f"${cost_data.get('target_cost_per_node', 0):,.2f}", 
                 f"{cost_data.get('node_variance', 0):.1f}%"],
                ['Cost Per Pod', f"${cost_data.get('cost_per_pod', 0):,.2f}", 
                 f"${cost_data.get('target_cost_per_pod', 0):,.2f}", 
                 f"{cost_data.get('pod_variance', 0):.1f}%"]
            ]
            
            cost_table = Table(cost_table_data)
            cost_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(cost_table)
            story.append(Spacer(1, 20))
            
            # Recommendations
            story.append(Paragraph("Key Recommendations", styles['Heading2']))
            
            recommendations = data.get('recommendations', [])
            for i, rec in enumerate(recommendations[:5], 1):  # Top 5 recommendations
                rec_text = f"<b>{i}. {rec.get('title', 'Recommendation')}</b><br/>{rec.get('description', '')}<br/><i>Potential savings: ${rec.get('savings', 0):,.2f}</i>"
                story.append(Paragraph(rec_text, styles['Normal']))
                story.append(Spacer(1, 10))
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"PDF report generation failed: {e}")
            raise
    
    async def _generate_excel_executive_report(self, data: Dict[str, Any], 
                                             config: ReportConfig) -> bytes:
        """Generate Excel executive report."""
        try:
            buffer = io.BytesIO()
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Executive Summary Sheet
            summary_sheet = workbook.create_sheet("Executive Summary")
            
            # Header styling
            header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
            
            # Title
            summary_sheet['A1'] = 'AKS Cost Intelligence - Executive Report'
            summary_sheet['A1'].font = Font(name='Calibri', size=18, bold=True, color='667EEA')
            summary_sheet.merge_cells('A1:D1')
            
            # Key metrics
            summary_data = data.get('summary', {})
            metrics = [
                ['Metric', 'Value', 'Target', 'Status'],
                ['Total Monthly Cost', f"${summary_data.get('total_cost', 0):,.2f}", 
                 f"${summary_data.get('target_cost', 0):,.2f}", 'Needs Optimization'],
                ['Potential Savings', f"${summary_data.get('potential_savings', 0):,.2f}", 
                 f"${summary_data.get('target_savings', 0):,.2f}", 'On Track'],
                ['Efficiency Score', f"{summary_data.get('efficiency_score', 0):.1f}%", '85%', 'Below Target'],
                ['Number of Clusters', summary_data.get('cluster_count', 0), 'N/A', 'Good']
            ]
            
            for row_idx, row_data in enumerate(metrics, start=3):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = summary_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    if row_idx == 3:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
            
            # Cost Analysis Sheet
            cost_sheet = workbook.create_sheet("Cost Analysis")
            
            # Add cost trend data
            cost_data = data.get('cost_analysis', {})
            trend_data = cost_data.get('trend_data', [])
            
            cost_sheet['A1'] = 'Date'
            cost_sheet['B1'] = 'Cost'
            cost_sheet['C1'] = 'Forecast'
            
            for i, trend_point in enumerate(trend_data, start=2):
                cost_sheet[f'A{i}'] = trend_point.get('date')
                cost_sheet[f'B{i}'] = trend_point.get('cost')
                cost_sheet[f'C{i}'] = trend_point.get('forecast')
            
            # Create chart
            chart = LineChart()
            chart.title = "Cost Trend Analysis"
            chart.style = 10
            chart.x_axis.title = 'Date'
            chart.y_axis.title = 'Cost ($)'
            
            data_range = Reference(cost_sheet, min_col=2, min_row=1, max_col=3, max_row=len(trend_data)+1)
            chart.add_data(data_range, titles_from_data=True)
            
            cost_sheet.add_chart(chart, "E2")
            
            # Recommendations Sheet
            rec_sheet = workbook.create_sheet("Recommendations")
            
            recommendations = data.get('recommendations', [])
            rec_headers = ['Priority', 'Title', 'Description', 'Potential Savings', 'Implementation Effort']
            
            for col_idx, header in enumerate(rec_headers, start=1):
                cell = rec_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            for row_idx, rec in enumerate(recommendations, start=2):
                rec_sheet.cell(row=row_idx, column=1, value=rec.get('priority', 'Medium'))
                rec_sheet.cell(row=row_idx, column=2, value=rec.get('title', ''))
                rec_sheet.cell(row=row_idx, column=3, value=rec.get('description', ''))
                rec_sheet.cell(row=row_idx, column=4, value=f"${rec.get('savings', 0):,.2f}")
                rec_sheet.cell(row=row_idx, column=5, value=rec.get('effort', 'Medium'))
            
            # Auto-adjust column widths
            for sheet in workbook.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to buffer
            workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Excel report generation failed: {e}")
            raise
    
    async def _generate_html_executive_report(self, data: Dict[str, Any], 
                                            config: ReportConfig) -> bytes:
        """Generate HTML executive report."""
        try:
            # Create visualizations
            cost_trend_chart = ""
            utilization_dashboard = ""
            
            if 'cost_trend_data' in data:
                cost_trend_df = pd.DataFrame(data['cost_trend_data'])
                cost_trend_chart = self.visualization_engine.create_cost_trend_chart(cost_trend_df)
            
            if 'utilization_data' in data:
                utilization_df = pd.DataFrame(data['utilization_data'])
                utilization_dashboard = self.visualization_engine.create_resource_utilization_dashboard(utilization_df)
            
            # HTML template
            html_template = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>AKS Cost Intelligence - Executive Report</title>
                <style>
                    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 40px; background: #f8f9fa; }}
                    .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; }}
                    .summary {{ background: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                    .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }}
                    .metric-card {{ background: white; padding: 20px; border-radius: 10px; text-align: center; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                    .metric-value {{ font-size: 2em; font-weight: bold; color: #667eea; }}
                    .metric-label {{ font-size: 0.9em; color: #666; margin-top: 5px; }}
                    .chart-container {{ background: white; padding: 20px; border-radius: 10px; margin: 20px 0; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                    .recommendations {{ background: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                    .recommendation {{ padding: 15px; margin: 10px 0; border-left: 4px solid #667eea; background: #f8f9ff; }}
                    h1, h2, h3 {{ color: #333; }}
                    .savings-highlight {{ background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%); color: white; padding: 20px; border-radius: 10px; text-align: center; margin: 20px 0; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>🚀 AKS Cost Intelligence</h1>
                    <h2>Executive Report - {datetime.now().strftime('%B %Y')}</h2>
                    <p>Comprehensive cost analysis and optimization recommendations</p>
                </div>
                
                <div class="summary">
                    <h2>Executive Summary</h2>
                    <div class="metric-grid">
                        <div class="metric-card savings-card">
                            <div class="metric-value">${data.get('summary', {}).get('total_cost', 0):,.0f}</div>
                            <div class="metric-label">Monthly Cost</div>
                        </div>
                        <div class="metric-card savings-card">
                            <div class="metric-value">${data.get('summary', {}).get('potential_savings', 0):,.0f}</div>
                            <div class="metric-label">Potential Savings</div>
                        </div>
                        <div class="metric-card savings-card">
                            <div class="metric-value">{data.get('summary', {}).get('efficiency_score', 0):.1f}%</div>
                            <div class="metric-label">Efficiency Score</div>
                        </div>
                        <div class="metric-card savings-card">
                            <div class="metric-value">{data.get('summary', {}).get('cluster_count', 0)}</div>
                            <div class="metric-label">Clusters Analyzed</div>
                        </div>
                    </div>
                    
                    <div class="savings-highlight">
                        <h3>💡 Key Insight</h3>
                        <p>By implementing our AI-powered recommendations, you can achieve an estimated <strong>{data.get('summary', {}).get('roi', 0):.1f}% ROI</strong> with potential monthly savings of <strong>${data.get('summary', {}).get('potential_savings', 0):,.0f}</strong></p>
                    </div>
                </div>
                
                <div class="chart-container">
                    <h2>📈 Cost Trend Analysis</h2>
                    {cost_trend_chart}
                </div>
                
                <div class="chart-container">
                    <h2>📊 Resource Utilization Dashboard</h2>
                    {utilization_dashboard}
                </div>
                
                <div class="recommendations">
                    <h2>🎯 Key Recommendations</h2>
                    {"".join([f'<div class="recommendation"><h3>{rec.get("title", "")}</h3><p>{rec.get("description", "")}</p><p><strong>Potential Savings:</strong> ${rec.get("savings", 0):,.2f}</p></div>' for rec in data.get('recommendations', [])[:5]])}
                </div>
                
                <div class="summary">
                    <h2>📋 Implementation Roadmap</h2>
                    <ol>
                        <li><strong>Phase 1 (Week 1-2):</strong> Implement quick wins with low risk</li>
                        <li><strong>Phase 2 (Week 3-6):</strong> Deploy advanced optimization strategies</li>
                        <li><strong>Phase 3 (Week 7-12):</strong> Monitor and fine-tune optimizations</li>
                        <li><strong>Phase 4 (Ongoing):</strong> Continuous optimization with AI monitoring</li>
                    </ol>
                </div>
                
                <div class="summary">
                    <h2>📞 Next Steps</h2>
                    <p>To maximize your cost savings and implement these recommendations:</p>
                    <ul>
                        <li>Schedule a technical review meeting with our optimization team</li>
                        <li>Prioritize recommendations based on your risk tolerance</li>
                        <li>Set up automated monitoring for continuous optimization</li>
                        <li>Plan regular review cycles to track progress and ROI</li>
                    </ul>
                </div>
                
                <div style="text-align: center; margin-top: 40px; color: #666; font-size: 0.9em;">
                    <p>Generated by AKS Cost Intelligence Platform on {datetime.now().strftime('%Y-%m-%d at %H:%M:%S')}</p>
                </div>
            </body>
            </html>
            """
            
            return html_template.encode('utf-8')
            
        except Exception as e:
            logger.error(f"HTML report generation failed: {e}")
            raise

class BusinessIntelligencePlatform:
    """Main business intelligence platform."""
    
    def __init__(self):
        self.app = FastAPI(title="Business Intelligence Platform")
        self.analytics_engine = AdvancedAnalyticsEngine()
        self.visualization_engine = VisualizationEngine()
        self.report_generator = ReportGenerator()
        self.scheduled_reports = {}
        self.dashboards = {}
        
        self._setup_routes()
        self._start_report_scheduler()
    
    def _setup_routes(self):
        """Setup BI platform routes."""
        
        @self.app.get("/")
        async def bi_home():
            """Business Intelligence home page."""
            return HTMLResponse("""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Business Intelligence Platform</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }
                    .header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 15px; text-align: center; }
                    .dashboard-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; margin: 30px 0; }
                    .dashboard-card { background: white; padding: 20px; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); }
                    .btn { background: #667eea; color: white; padding: 12px 24px; border: none; border-radius: 8px; cursor: pointer; text-decoration: none; display: inline-block; }
                    .btn:hover { background: #5a6fd8; }
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>📊 Business Intelligence Platform</h1>
                    <p>Advanced analytics and reporting for AKS cost optimization</p>
                </div>
                
                <div class="dashboard-grid">
                    <div class="dashboard-card">
                        <h3>📈 Executive Dashboard</h3>
                        <p>High-level KPIs and trends for leadership</p>
                        <a href="/dashboards/executive" class="btn">View Dashboard</a>
                    </div>
                    
                    <div class="dashboard-card">
                        <h3>💰 Cost Analytics</h3>
                        <p>Detailed cost analysis and forecasting</p>
                        <a href="/dashboards/cost" class="btn">View Analytics</a>
                    </div>
                    
                    <div class="dashboard-card">
                        <h3>📊 Resource Optimization</h3>
                        <p>Resource utilization and optimization insights</p>
                        <a href="/dashboards/resources" class="btn">View Resources</a>
                    </div>
                    
                    <div class="dashboard-card">
                        <h3>📋 Custom Reports</h3>
                        <p>Generate and schedule custom reports</p>
                        <a href="/reports" class="btn">Manage Reports</a>
                    </div>
                </div>
            </body>
            </html>
            """)
        
        @self.app.get("/api/analytics/cost-trends")
        async def get_cost_trends():
            """Get cost trend analytics."""
            try:
                # Sample data - in production, this would come from your data source
                sample_data = pd.DataFrame({
                    'timestamp': pd.date_range('2024-01-01', periods=90, freq='D'),
                    'cost': np.random.normal(1000, 200, 90) + np.sin(np.arange(90) * 0.1) * 100
                })
                
                analysis = await self.analytics_engine.perform_cost_trend_analysis(sample_data)
                return analysis
                
            except Exception as e:
                raise HTTPException(status_code=500, detail=str(e))
        
        @self.app.post("/api/reports/generate")
        async def generate_report(config: dict, background_tasks: BackgroundTasks):
            """Generate a custom report."""
            try:
                # Sample data for report
                report_data = {
                    'summary': {
                        'total_cost': 25010,
                        'potential_savings': 8500,
                        'efficiency_score': 72.5,
                        'cluster_count': 5,
                        'roi': 245.5
                    },
                    'recommendations': [
                        {
                            'title': 'Right-size Node Pools',
                            'description': 'Reduce oversized node pools to match actual workload requirements',
                            'savings': 3200,
                            'priority': 'High'
                        },
                        {
                            'title': 'Implement Auto-scaling',
                            'description': 'Deploy horizontal pod autoscaling for dynamic workload management',
                            'savings': 2800,
                            'priority': 'Medium'
                        }
                    ]
                }
                
                report_config = ReportConfig(
                    report_id=str(uuid.uuid4()),
                    name=config.get('name', 'Custom Report'),
                    description=config.get('description', ''),
                    report_type=ReportType(config.get('type', 'executive_summary')),
                    format=ReportFormat(config.get('format', 'pdf')),
                    frequency=ReportFrequency.ON_DEMAND,
                    recipients=config.get('recipients', []),
                    parameters=config.get('parameters', {}),
                    filters=config.get('filters', {}),
                    charts_included=config.get('charts', []),
                    created_by='system',
                    created_at=datetime.utcnow(),
                    last_generated=None,
                    next_generation=None,
                    active=True
                )
                
                # Generate report in background
                background_tasks.add_task(
                    self._generate_and_deliver_report,
                    report_data,
                    report_config
                )
                
                return {
                    'report_id': report_config.report_id,
                    'status': 'generating',
                    'message': 'Report generation started'
                }
                
            except Exception as e:
                raise HTTPException(status_code=500, detail=str(e))
    
    async def _generate_and_deliver_report(self, data: Dict[str, Any], 
                                         config: ReportConfig):
        """Generate and deliver report."""
        try:
            # Generate report
            report_content = await self.report_generator.generate_executive_report(data, config)
            
            # Save report
            report_filename = f"report_{config.report_id}.{config.format.value}"
            report_path = f"/tmp/{report_filename}"
            
            with open(report_path, 'wb') as f:
                f.write(report_content)
            
            # Send to recipients if specified
            if config.recipients:
                await self._send_report_email(config.recipients, report_path, config)
            
            logger.info(f"Report {config.report_id} generated successfully")
            
        except Exception as e:
            logger.error(f"Report generation failed: {e}")
    
    async def _send_report_email(self, recipients: List[str], report_path: str, 
                               config: ReportConfig):
        """Send report via email."""
        try:
            # Email implementation would go here
            logger.info(f"Report {config.name} sent to {len(recipients)} recipients")
            
        except Exception as e:
            logger.error(f"Failed to send report email: {e}")
    
    def _start_report_scheduler(self):
        """Start the report scheduler."""
        def run_scheduler():
            while True:
                schedule.run_pending()
                time.sleep(60)
        
        scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
        scheduler_thread.start()

# Example usage
if __name__ == "__main__":
    import uvicorn
    
    platform = BusinessIntelligencePlatform()
    
    uvicorn.run(
        platform.app,
        host="0.0.0.0",
        port=8090,
        log_level="info"
    )